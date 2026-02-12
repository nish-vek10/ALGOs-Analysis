# src/scripts/02_quant_report_processed_news_filter.py
"""
Quant-style report from a PROCESSED closed-trades CSV (CLOSED POSITIONS ONLY),
but with a NEWS blackout filter applied.

Goal:
- Remove any trades whose OpenTime falls within +/- window_minutes of any event.
- Compute the same report (metrics/plots/heatmaps) on the remaining trades.
- Include additional Excel sheets:
  - Redacted Summary (metrics on removed trades)
  - Redacted Trades (table of removed trades for review)

Inputs:
- PROCESSED_CSV (same as your current processed script)
- NEWS_JSON (config/news_events.json)

Output:
  output/runs/<RUN_ID>/
    meta/summary.json
    report/summary.xlsx
    tables/trades_with_equity.csv
    tables/redacted_trades.csv
    figures/equity_drawdown_pct.png
    figures/heatmaps/heatmap_*.png
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
PROCESSED_CSV = Path("data/processed/ReportHistory-52651106__acct-52651106__offset-2h.csv")
NEWS_JSON = Path("config/news_events.json")

START_EQUITY = 9400.0
MT_TIME_FMT = "%Y.%m.%d %H:%M:%S"

PNL_COL = "Profit"
DROP_ZERO_PNL_ROWS = False
ZERO_EPS = 1e-9

STRATEGY_COL = "Comment"
STRATEGY_MAP = {
    "Inv_ATRtrail_pen": "ATR-Trailing STRAT",
    "Inv_Chandelier_p": "Chandelier STRAT",
}
STRATEGY_ORDER = ["ATR-Trailing STRAT", "Chandelier STRAT"]

# The columns you want shown for redacted trades sheet (we’ll map what exists)
REDACTED_COLS_PREFERRED = [
    "OpenTime", "Position", "Symbol", "Type", "Volume", "OpenPrice", "S / L", "T / P",
    "CloseTime", "ClosePrice", "Commission", "Swap", "Profit", "Comment"
]


# =========================
# PATHS
# =========================
def project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def extract_account_id_from_filename(p: Path) -> str:
    m = re.search(r"(\d{5,})", p.stem)
    return m.group(1) if m else "unknown"


def safe_slug(s: str, max_len: int = 90) -> str:
    s = re.sub(r"[^A-Za-z0-9_\-]+", "_", s).strip("_")
    return s[:max_len] if len(s) > max_len else s


def utc_run_id(acct: str, stem: str) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") + "Z"
    return f"{ts}__acct-{safe_slug(acct, 40)}__NEWS_FILTER__{safe_slug(stem, 80)}"


def ensure_run_dirs(run_dir: Path) -> Dict[str, Path]:
    dirs = {
        "run": run_dir,
        "meta": run_dir / "meta",
        "tables": run_dir / "tables",
        "figures": run_dir / "figures",
        "heatmaps": run_dir / "figures" / "heatmaps",
        "report": run_dir / "report",
    }
    for p in dirs.values():
        p.mkdir(parents=True, exist_ok=True)
    return dirs


# =========================
# HELPERS
# =========================
def json_safe(o):
    import math
    import numpy as np
    import pandas as pd
    from datetime import datetime, date

    if isinstance(o, (pd.Timestamp,)):
        return None if pd.isna(o) else o.isoformat()

    if isinstance(o, (datetime, date)):
        return o.isoformat()

    if isinstance(o, (np.integer,)):
        return int(o)
    if isinstance(o, (np.floating,)):
        v = float(o)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(o, (np.bool_,)):
        return bool(o)

    if isinstance(o, np.ndarray):
        return o.tolist()

    try:
        if o is None:
            return None
        if isinstance(o, float) and (math.isnan(o) or math.isinf(o)):
            return None
    except Exception:
        pass

    return str(o)


def coerce_numeric(series: pd.Series) -> pd.Series:
    s = series.astype(str)
    s = s.str.replace("\u2212", "-", regex=False).str.replace("\xa0", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    s = s.str.replace(",", "", regex=False)
    s = s.str.replace(r"[^0-9\.\-]", "", regex=True)
    return pd.to_numeric(s, errors="coerce")


def parse_mt_time(series: pd.Series) -> pd.Series:
    dt = pd.to_datetime(series, format=MT_TIME_FMT, errors="coerce")
    if dt.notna().sum() == 0:
        dt = pd.to_datetime(series, errors="coerce")
    return dt


# =========================
# COLUMN DETECTION
# =========================
def find_time_column(df: pd.DataFrame, kind: str) -> Optional[str]:
    if kind not in {"open", "close"}:
        raise ValueError("kind must be 'open' or 'close'")

    candidates = (
        ["OpenTime", "Open Time", "open_time", "opentime"] if kind == "open"
        else ["CloseTime", "Close Time", "close_time", "closetime"]
    )

    cols_lower = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in cols_lower:
            return cols_lower[c.lower()]

    for c in df.columns:
        cl = c.lower()
        if kind == "open" and ("open" in cl and "time" in cl):
            return c
        if kind == "close" and ("close" in cl and "time" in cl):
            return c

    return None


def normalize_strategy(raw: object) -> str:
    s = "" if raw is None else str(raw).strip()
    return STRATEGY_MAP.get(s, s if s else "Unknown STRAT")


# =========================
# METRICS
# =========================
def max_drawdown_pct(equity: np.ndarray, start_equity: float) -> Tuple[float, int, int]:
    if len(equity) == 0 or start_equity == 0:
        return 0.0, -1, -1
    peaks = np.maximum.accumulate(equity)
    dd_cash = peaks - equity
    trough_idx = int(np.argmax(dd_cash))
    peak_idx = int(np.argmax(peaks[:trough_idx + 1])) if trough_idx >= 0 else 0
    max_dd_cash = float(dd_cash[trough_idx]) if len(dd_cash) else 0.0
    return float((max_dd_cash / start_equity) * 100.0), peak_idx, trough_idx


def sharpe_ratio_per_trade(pnl: np.ndarray, start_equity: float) -> float:
    if start_equity <= 0 or len(pnl) < 2:
        return 0.0
    rets = pnl / float(start_equity)
    mu = float(np.mean(rets))
    sd = float(np.std(rets, ddof=1))
    return 0.0 if sd == 0.0 else (mu / sd)


def compute_metrics(df_sub: pd.DataFrame, start_equity: float) -> dict:
    total_trades = int(len(df_sub))
    pnl = df_sub["_pnl"].to_numpy(dtype=float) if total_trades else np.array([], dtype=float)

    winners = int((df_sub["_pnl"] > 0).sum())
    losers = int((df_sub["_pnl"] < 0).sum())
    breakeven = int((df_sub["_pnl"] == 0.0).sum())

    win_rate = (winners / total_trades) * 100.0 if total_trades else 0.0

    cum_pnl = np.cumsum(pnl) if total_trades else np.array([], dtype=float)
    equity = start_equity + cum_pnl if total_trades else np.array([start_equity], dtype=float)

    total_pnl = float(cum_pnl[-1]) if total_trades else 0.0
    end_equity = float(equity[-1]) if total_trades else float(start_equity)
    net_pct = ((end_equity / float(start_equity)) - 1.0) * 100.0 if start_equity else 0.0

    max_dd_pct_pos, _, _ = max_drawdown_pct(equity, start_equity)

    gross_win = float(df_sub.loc[df_sub["_pnl"] > 0, "_pnl"].sum()) if winners else 0.0
    gross_loss = float(df_sub.loc[df_sub["_pnl"] < 0, "_pnl"].sum()) if losers else 0.0
    profit_factor = float(gross_win / abs(gross_loss)) if losers else float("inf")

    expectancy = float(total_pnl / total_trades) if total_trades else 0.0

    best_trade = float(df_sub["_pnl"].max()) if total_trades else 0.0
    worst_trade = float(df_sub["_pnl"].min()) if total_trades else 0.0

    avg_win = float(df_sub.loc[df_sub["_pnl"] > 0, "_pnl"].mean()) if winners else 0.0
    avg_loss = float(df_sub.loc[df_sub["_pnl"] < 0, "_pnl"].mean()) if losers else 0.0

    hold = df_sub["_hold_minutes"].dropna()
    avg_hold = float(hold.mean()) if len(hold) else None
    med_hold = float(hold.median()) if len(hold) else None
    p95_hold = float(hold.quantile(0.95)) if len(hold) else None

    sharpe = float(sharpe_ratio_per_trade(pnl, start_equity)) if total_trades else 0.0
    recovery = float(net_pct / max_dd_pct_pos) if max_dd_pct_pos not in (0.0, -0.0) else 0.0

    return {
        "totalTrades": total_trades,
        "winners": winners,
        "losers": losers,
        "breakeven": breakeven,
        "winRate_pct": win_rate,

        "startEquity": float(start_equity),
        "endEquity": end_equity,
        "pnl_cash": total_pnl,
        "netPct": net_pct,

        "maxDD_pct": max_dd_pct_pos,
        "grossWin": gross_win,
        "grossLoss": gross_loss,
        "profitFactor": profit_factor,

        "expectancy_per_trade": expectancy,
        "bestTrade": best_trade,
        "worstTrade": worst_trade,
        "avgWin": avg_win,
        "avgLoss": avg_loss,

        "avgHoldMin": avg_hold,
        "medianHoldMin": med_hold,
        "p95HoldMin": p95_hold,

        "sharpe_ratio": sharpe,
        "recovery_factor": recovery,
    }


# =========================
# NEWS FILTER
# =========================
def load_news_config(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"News JSON not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def parse_news_events(cfg: dict) -> Tuple[List[dict], int]:
    window_minutes = int(cfg.get("window_minutes", 7))
    events = cfg.get("events", [])
    parsed = []
    for e in events:
        name = str(e.get("name", "NEWS")).strip() or "NEWS"
        dt_raw = str(e.get("datetime", "")).strip()
        if not dt_raw:
            continue
        # expected: "YYYY-MM-DD HH:MM"
        dt = pd.to_datetime(dt_raw, errors="coerce")
        if pd.isna(dt):
            continue
        parsed.append({"name": name, "dt": dt.to_pydatetime()})
    return parsed, window_minutes


def mask_trades_in_news_window(open_times: pd.Series, events: List[dict], window_minutes: int) -> pd.Series:
    """
    Vector mask: True if open_time within +/- window of any event.
    """
    ot = pd.to_datetime(open_times)
    mask = pd.Series(False, index=ot.index)

    if not len(events):
        return mask

    delta = pd.Timedelta(minutes=window_minutes)
    for e in events:
        t0 = pd.Timestamp(e["dt"])
        mask |= (ot >= (t0 - delta)) & (ot <= (t0 + delta))

    return mask


# =========================
# PLOTS (same style as your processed)
# =========================
def plot_equity_dd(times: pd.Series, equity: np.ndarray, start_equity: float, out_path: Path, title: str) -> None:
    import matplotlib.dates as mdates

    times = pd.to_datetime(times)
    equity = np.asarray(equity, dtype=float)

    eq_pct = ((equity / float(start_equity)) - 1.0) * 100.0 if start_equity else np.zeros_like(equity)

    peaks = np.maximum.accumulate(equity)
    dd_cash = peaks - equity
    dd_pct = (dd_cash / float(start_equity)) * 100.0 if start_equity else np.zeros_like(dd_cash)
    dd_pct = -dd_pct

    fig = plt.figure(figsize=(16, 7.5), facecolor="#F5F7FB", constrained_layout=True)
    gs = fig.add_gridspec(2, 1, height_ratios=[2.5, 1.15])

    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1], sharex=ax1)

    def style_card(ax):
        ax.set_facecolor("white")
        ax.grid(True, which="major", alpha=0.18, linewidth=1)
        ax.grid(True, which="minor", alpha=0.08, linewidth=0.8)
        for side in ["top", "right"]:
            ax.spines[side].set_visible(False)
        for side in ["left", "bottom"]:
            ax.spines[side].set_alpha(0.25)
        ax.tick_params(axis="both", labelsize=10)

    style_card(ax1)
    style_card(ax2)

    ax1.set_title(title, fontsize=16, fontweight="bold", pad=12)

    ax1.plot(times, eq_pct, linewidth=2.6, alpha=0.95)
    ax1.axhline(0, linewidth=1.0, alpha=0.35)
    ax1.fill_between(times, eq_pct, 0, where=(eq_pct >= 0), alpha=0.18, interpolate=True)
    ax1.fill_between(times, eq_pct, 0, where=(eq_pct < 0), color="red", alpha=0.10, interpolate=True)
    ax1.set_ylabel("Equity (%)", fontsize=12)
    ax1.tick_params(axis="x", labelbottom=False)

    ax2.plot(times, dd_pct, linewidth=2.2, color="red", alpha=0.95)
    ax2.fill_between(times, dd_pct, 0, color="red", alpha=0.12)
    ax2.axhline(0, linewidth=1.0, alpha=0.35)
    ax2.set_ylabel("Drawdown (%)", fontsize=12)
    ax2.set_xlabel("Time", fontsize=12)

    dd_min = float(np.nanmin(dd_pct)) if len(dd_pct) else 0.0
    ax2.set_ylim(min(dd_min * 1.15, -0.5), 0.0)

    ax2.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=10))
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    ax2.xaxis.set_minor_locator(mdates.AutoDateLocator(minticks=12, maxticks=20))
    fig.autofmt_xdate(rotation=0)

    if len(eq_pct):
        ax1.annotate(
            f"End: {eq_pct[-1]:.2f}%",
            xy=(times.iloc[-1], eq_pct[-1]),
            xytext=(-12, 10),
            textcoords="offset points",
            ha="right",
            fontsize=10,
            alpha=0.85,
        )

    fig.savefig(out_path, dpi=240, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_equity_dd_multi(
    times: pd.Series,
    curves: Dict[str, np.ndarray],
    start_equity: float,
    out_path: Path,
    title: str,
) -> None:
    """
    Same styling as plot_equity_dd, but overlays multiple equity curves (per strategy)
    and plots each strategy's drawdown on the lower panel.

    curves: {label -> equity_array_in_cash}
    """

    import matplotlib.dates as mdates

    times = pd.to_datetime(times)

    fig = plt.figure(figsize=(16, 7.5), facecolor="#F5F7FB", constrained_layout=True)
    gs = fig.add_gridspec(2, 1, height_ratios=[2.5, 1.15])
    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1], sharex=ax1)

    def style_card(ax):
        ax.set_facecolor("white")
        ax.grid(True, which="major", alpha=0.18, linewidth=1)
        ax.grid(True, which="minor", alpha=0.08, linewidth=0.8)
        for side in ["top", "right"]:
            ax.spines[side].set_visible(False)
        for side in ["left", "bottom"]:
            ax.spines[side].set_alpha(0.25)
        ax.tick_params(axis="both", labelsize=10)

    style_card(ax1)
    style_card(ax2)

    ax1.set_title(title, fontsize=16, fontweight="bold", pad=12)

    # baseline
    ax1.axhline(0, linewidth=1.0, alpha=0.35)
    ax2.axhline(0, linewidth=1.0, alpha=0.35)

    # plot each strategy
    dd_mins = []
    for label, eq in curves.items():
        eq = np.asarray(eq, dtype=float)

        # equity %
        eq_pct = ((eq / float(start_equity)) - 1.0) * 100.0 if start_equity else np.zeros_like(eq)
        ax1.plot(times, eq_pct, linewidth=2.4, alpha=0.95, label=label)

        # drawdown % (negative)
        peaks = np.maximum.accumulate(eq)
        dd_cash = peaks - eq
        dd_pct = (dd_cash / float(start_equity)) * 100.0 if start_equity else np.zeros_like(dd_cash)
        dd_pct = -dd_pct
        ax2.plot(times, dd_pct, linewidth=2.0, alpha=0.9, label=label)

        if len(dd_pct):
            dd_mins.append(float(np.nanmin(dd_pct)))

    # labels + legends
    ax1.set_ylabel("Equity (%)", fontsize=12)
    ax1.tick_params(axis="x", labelbottom=False)

    ax2.set_ylabel("Drawdown (%)", fontsize=12)
    ax2.set_xlabel("Time", fontsize=12)

    ax1.legend(loc="upper left", frameon=False, fontsize=10)
    ax2.legend(loc="lower left", frameon=False, fontsize=10)

    # dd limits padding
    dd_min = min(dd_mins) if dd_mins else 0.0
    ax2.set_ylim(min(dd_min * 1.15, -0.5), 0.0)

    # dates
    ax2.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=10))
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    ax2.xaxis.set_minor_locator(mdates.AutoDateLocator(minticks=12, maxticks=20))
    fig.autofmt_xdate(rotation=0)

    fig.savefig(out_path, dpi=240, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_heatmap(mat: pd.DataFrame, out_path: Path, title: str, fmt: str, higher_is_better: bool) -> None:
    day_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    mat = mat.reindex(day_order)
    mat = mat.reindex(columns=list(range(24)))

    values = mat.to_numpy(dtype=float)
    cmap = plt.cm.RdYlGn if higher_is_better else plt.cm.RdYlGn_r

    fig, ax = plt.subplots(figsize=(16, 5.5), constrained_layout=True)
    im = ax.imshow(values, aspect="auto", cmap=cmap)

    ax.set_title(title)
    ax.set_xlabel("Hour of day (chart time)")
    ax.set_ylabel("Day of week")

    ax.set_xticks(np.arange(24))
    ax.set_xticklabels([str(h) for h in range(24)])
    ax.set_yticks(np.arange(len(day_order)))
    ax.set_yticklabels(day_order)

    cbar = fig.colorbar(im, ax=ax, fraction=0.02, pad=0.02)
    cbar.ax.tick_params(labelsize=9)

    for i in range(values.shape[0]):
        for j in range(values.shape[1]):
            v = values[i, j]
            if np.isfinite(v):
                ax.text(j, i, format(v, fmt), ha="center", va="center", fontsize=7)

    fig.savefig(out_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def bin_max_dd_pct(sub: pd.DataFrame, start_equity: float) -> float:
    if len(sub) == 0 or start_equity == 0:
        return np.nan
    sub = sub.sort_values("_open_dt")
    cum = np.cumsum(sub["_pnl"].to_numpy(dtype=float))
    peaks = np.maximum.accumulate(cum)
    dd = peaks - cum
    return float(np.max(dd) / float(start_equity) * 100.0) if len(dd) else 0.0


# =========================
# EXCEL (Summary + Redacted)
# =========================
def autosize_columns(ws, max_width: int = 46) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))


def write_summary_sheet(ws, rows: list[dict], title: str) -> None:
    ws.title = title

    headers = [
        "Strategy",
        "totalTrades", "winners", "losers", "breakeven",
        "startEquity", "endEquity",
        "winRate_pct",
        "pnl_cash", "netPct",
        "maxDD_pct",
        "profitFactor",
        "expectancy_per_trade",
        "bestTrade", "worstTrade",
        "avgWin", "avgLoss",
        "grossWin", "grossLoss",
        "sharpe_ratio", "recovery_factor",
        "avgHoldMin", "medianHoldMin", "p95HoldMin",
    ]

    ws.append(headers)

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="F2F4F8")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = fill

    for r in rows:
        ws.append([r.get(h) for h in headers])

    autosize_columns(ws)


def write_redacted_trades_sheet(wb: Workbook, df_redacted: pd.DataFrame) -> None:
    ws = wb.create_sheet("Redacted Trades")

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="F2F4F8")

    # Choose best column mapping that exists in the processed file
    cols = [c for c in REDACTED_COLS_PREFERRED if c in df_redacted.columns]

    # Always include computed helpers if present
    for c in ["_open_dt", "_close_dt", "_pnl", "_strategy", "_redact_reason"]:
        if c in df_redacted.columns and c not in cols:
            cols.append(c)

    ws.append(cols)
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j)
        cell.font = header_font
        cell.alignment = center
        cell.fill = fill

    for _, row in df_redacted[cols].iterrows():
        ws.append([row.get(c) for c in cols])

    autosize_columns(ws, max_width=60)


def write_summary_xlsx(path: Path, rows_main: list[dict], rows_redacted: list[dict], df_redacted: pd.DataFrame) -> None:
    wb = Workbook()

    ws_main = wb.active
    write_summary_sheet(ws_main, rows_main, title="Summary (Filtered)")

    ws_red_sum = wb.create_sheet("Redacted Summary")
    write_summary_sheet(ws_red_sum, rows_redacted, title="Redacted Summary")

    write_redacted_trades_sheet(wb, df_redacted)

    wb.save(path)


# =========================
# MAIN
# =========================
def main() -> None:
    root = project_root()
    processed_path = (root / PROCESSED_CSV).resolve()
    if not processed_path.exists():
        raise FileNotFoundError(f"Processed CSV not found: {processed_path}")

    news_path = (root / NEWS_JSON).resolve()
    cfg = load_news_config(news_path)
    events, window_minutes = parse_news_events(cfg)

    df = pd.read_csv(processed_path)

    open_col = find_time_column(df, "open")
    close_col = find_time_column(df, "close")
    if open_col is None:
        raise ValueError(f"Could not find an OpenTime column. Columns={list(df.columns)}")
    if PNL_COL not in df.columns:
        raise ValueError(f"Expected per-trade PnL column '{PNL_COL}' not found. Columns={list(df.columns)}")

    df["_open_dt"] = parse_mt_time(df[open_col])
    df["_close_dt"] = parse_mt_time(df[close_col]) if close_col is not None else pd.NaT

    # closed-only
    df = df[df["_open_dt"].notna()].copy()
    if close_col is not None:
        df = df[df["_close_dt"].notna()].copy()

    df["_pnl"] = coerce_numeric(df[PNL_COL]).fillna(0.0)

    # strategy
    if STRATEGY_COL in df.columns:
        df["_strategy"] = df[STRATEGY_COL].apply(normalize_strategy)
    else:
        df["_strategy"] = "Unknown STRAT"

    if DROP_ZERO_PNL_ROWS:
        df = df[df["_pnl"].abs() > ZERO_EPS].copy()

    df.sort_values("_open_dt", inplace=True)
    df.reset_index(drop=True, inplace=True)

    df["_hold_minutes"] = (df["_close_dt"] - df["_open_dt"]).dt.total_seconds() / 60.0 if close_col is not None else np.nan

    acct = extract_account_id_from_filename(processed_path)
    run_id = utc_run_id(acct=acct, stem=processed_path.stem)
    run_dir = (root / "output" / "runs" / run_id).resolve()
    dirs = ensure_run_dirs(run_dir)

    # -------------------------
    # Apply news blackout filter (by OpenTime)
    # -------------------------
    mask_news = mask_trades_in_news_window(df["_open_dt"], events, window_minutes)

    df_redacted = df[mask_news].copy()
    df_kept = df[~mask_news].copy()

    # Add detailed event-match columns for proof (only after df_redacted exists)
    df_redacted["_redact_reason"] = ""
    df_redacted["_news_event"] = ""
    df_redacted["_news_time"] = ""
    df_redacted["_mins_from_event"] = np.nan

    if len(df_redacted) and len(events):
        delta = pd.Timedelta(minutes=window_minutes)
        # For speed: prebuild list of (name, Timestamp)
        ev = [(e["name"], pd.Timestamp(e["dt"])) for e in events]

        for i in df_redacted.index:
            ot = pd.Timestamp(df_redacted.loc[i, "_open_dt"])
            for name, t0 in ev:
                if (ot >= t0 - delta) and (ot <= t0 + delta):
                    df_redacted.loc[i, "_redact_reason"] = f"OpenTime within ±{window_minutes}m"
                    df_redacted.loc[i, "_news_event"] = name
                    df_redacted.loc[i, "_news_time"] = t0.strftime("%Y-%m-%d %H:%M")
                    df_redacted.loc[i, "_mins_from_event"] = (ot - t0).total_seconds() / 60.0
                    break

    # Save redacted trades CSV
    df_redacted.to_csv(dirs["tables"] / "redacted_trades.csv", index=False)

    print(f"[NEWS] events={len(events)} window=±{window_minutes}m redacted={len(df_redacted)} kept={len(df_kept)}")

    # -------------------------
    # Compute report on kept trades
    # -------------------------
    pnl = df_kept["_pnl"].to_numpy(dtype=float)
    cum_pnl = np.cumsum(pnl) if len(pnl) else np.array([], dtype=float)
    equity = START_EQUITY + cum_pnl if len(pnl) else np.array([START_EQUITY], dtype=float)

    peaks = np.maximum.accumulate(equity)
    dd_cash = peaks - equity
    dd_pct = (dd_cash / float(START_EQUITY)) * 100.0 if START_EQUITY else np.zeros_like(dd_cash)
    dd_pct = -dd_pct

    overall_metrics = compute_metrics(df_kept, START_EQUITY)
    st = {
        "max_win_streak_n": int((df_kept["_pnl"] > 0).astype(int).groupby((df_kept["_pnl"] <= 0).cumsum()).sum().max()) if len(df_kept) else 0,
        "max_loss_streak_n": int((df_kept["_pnl"] < 0).astype(int).groupby((df_kept["_pnl"] >= 0).cumsum()).sum().max()) if len(df_kept) else 0,
        # keep cash streaks via your existing logic? (optional)
    }
    max_dd_pct_pos, peak_i, trough_i = max_drawdown_pct(equity, START_EQUITY)

    # Per-strategy (kept)
    by_strategy = {}
    for strat_name, df_s in df_kept.groupby("_strategy", dropna=False):
        by_strategy[strat_name] = compute_metrics(df_s, START_EQUITY)

    # -------------------------
    # Redacted summary (same metric set, but calculated on removed trades only)
    # -------------------------
    red_overall_metrics = compute_metrics(df_redacted, START_EQUITY) if len(df_redacted) else compute_metrics(df_redacted, START_EQUITY)

    red_by_strategy = {}
    for strat_name, df_s in df_redacted.groupby("_strategy", dropna=False):
        red_by_strategy[strat_name] = compute_metrics(df_s, START_EQUITY)

    # -------------------------
    # Write summary.json
    # -------------------------
    summary = {
        "input": {
            "processed_csv": str(processed_path),
            "news_json": str(news_path),
            "window_minutes": window_minutes,
            "events_n": len(events),
            "event_names": [e["name"] for e in events],
            "start_equity": START_EQUITY,
            "open_time_col": open_col,
            "close_time_col": close_col,
            "pnl_col": PNL_COL,
        },
        "filter": {
            "kept_trades": int(len(df_kept)),
            "redacted_trades": int(len(df_redacted)),
            "rule": "Drop trades if OpenTime within ±window_minutes of any news event"
        },
        "totals_filtered": overall_metrics,
        "by_strategy_filtered": by_strategy,
        "totals_redacted": red_overall_metrics,
        "by_strategy_redacted": red_by_strategy,
        "max_dd_window_filtered": {
            "peak_index": int(peak_i),
            "trough_index": int(trough_i),
            "peak_time_open": str(df_kept["_open_dt"].iloc[peak_i]) if peak_i >= 0 and len(df_kept) else None,
            "trough_time_open": str(df_kept["_open_dt"].iloc[trough_i]) if trough_i >= 0 and len(df_kept) else None,
        },
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "run_dir": str(run_dir),
    }

    (dirs["meta"] / "summary.json").write_text(
        json.dumps(summary, indent=2, default=json_safe),
        encoding="utf-8"
    )

    # -------------------------
    # XLSX: Summary + Redacted
    # -------------------------
    rows_main = []
    for strat in STRATEGY_ORDER:
        df_s = df_kept[df_kept["_strategy"] == strat]
        m = compute_metrics(df_s, START_EQUITY)
        m["Strategy"] = strat
        rows_main.append(m)

    overall_row = overall_metrics.copy()
    overall_row["Strategy"] = "OVERALL (PORTFOLIO)"
    rows_main.append(overall_row)

    rows_red = []
    for strat in STRATEGY_ORDER:
        df_s = df_redacted[df_redacted["_strategy"] == strat]
        m = compute_metrics(df_s, START_EQUITY) if len(df_s) else compute_metrics(df_s, START_EQUITY)
        m["Strategy"] = strat
        rows_red.append(m)

    red_overall_row = red_overall_metrics.copy()
    red_overall_row["Strategy"] = "OVERALL (PORTFOLIO)"
    rows_red.append(red_overall_row)

    xlsx_path = dirs["report"] / "summary.xlsx"
    write_summary_xlsx(xlsx_path, rows_main, rows_red, df_redacted)

    # -------------------------
    # trades_with_equity.csv (filtered trades)
    # -------------------------
    df_out = df_kept.copy()
    df_out["CumPnL"] = np.cumsum(df_out["_pnl"].to_numpy(dtype=float)) if len(df_out) else []
    df_out["Equity"] = START_EQUITY + (df_out["CumPnL"].to_numpy(dtype=float) if len(df_out) else 0.0)
    peaks2 = np.maximum.accumulate(df_out["Equity"].to_numpy(dtype=float)) if len(df_out) else np.array([START_EQUITY])
    dd_cash2 = peaks2 - df_out["Equity"].to_numpy(dtype=float) if len(df_out) else np.array([0.0])
    df_out["Drawdown_pct"] = -(dd_cash2 / float(START_EQUITY) * 100.0) if (START_EQUITY and len(df_out)) else 0.0
    df_out.to_csv(dirs["tables"] / "trades_with_equity.csv", index=False)

    # -------------------------
    # Plot equity + drawdown (filtered)
    # -------------------------
    plot_equity_dd(
        times=df_kept["_open_dt"] if len(df_kept) else pd.Series([], dtype="datetime64[ns]"),
        equity=equity,
        start_equity=START_EQUITY,
        out_path=dirs["figures"] / "equity_drawdown_pct.png",
        title=f"Equity Curve & Drawdown (Filtered — No trades ±{window_minutes}m news) — acct {acct}",
    )

    # -------------------------
    # Per-strategy equity + drawdown (Filtered, two lines)
    # -------------------------
    if len(df_kept):
        timeline = pd.to_datetime(df_kept["_open_dt"]).dropna().sort_values().unique()
        timeline = pd.Series(timeline)

        curves_ff = {}
        for strat in STRATEGY_ORDER:
            df_s = df_kept[df_kept["_strategy"] == strat].sort_values("_open_dt")
            if len(df_s) == 0:
                continue
            t_s = pd.to_datetime(df_s["_open_dt"])
            eq_s = START_EQUITY + np.cumsum(df_s["_pnl"].to_numpy(dtype=float))
            s = pd.Series(eq_s, index=t_s)

            s2 = s.reindex(timeline, method="ffill")
            s2 = s2.fillna(START_EQUITY)
            curves_ff[strat] = s2.to_numpy(dtype=float)

        if len(curves_ff):
            plot_equity_dd_multi(
                times=timeline,
                curves=curves_ff,
                start_equity=START_EQUITY,
                out_path=dirs["figures"] / "equity_drawdown_pct_by_strategy.png",
                title=f"Equity Curve & Drawdown (Filtered — By Strategy, No trades ±{window_minutes}m news) — acct {acct}",
            )

    # -------------------------
    # Heatmaps (filtered) by OPEN time
    # -------------------------
    if len(df_kept):
        df_kept["_dow"] = df_kept["_open_dt"].dt.day_name().str[:3]
        df_kept["_hour"] = df_kept["_open_dt"].dt.hour
        df_kept["_is_win"] = (df_kept["_pnl"] > 0).astype(int)

        trade_count = df_kept.pivot_table(index="_dow", columns="_hour", values="_pnl", aggfunc="count")
        win_rate_mat = df_kept.pivot_table(index="_dow", columns="_hour", values="_is_win", aggfunc="mean") * 100.0
        pnl_sum = df_kept.pivot_table(index="_dow", columns="_hour", values="_pnl", aggfunc="sum")
        netpct_mat = (pnl_sum / float(START_EQUITY) * 100.0) if START_EQUITY else pnl_sum * 0.0

        parts = []
        for (dow, hour), sub in df_kept[["_dow", "_hour", "_open_dt", "_pnl"]].groupby(["_dow", "_hour"], sort=False):
            parts.append((dow, hour, bin_max_dd_pct(sub, START_EQUITY)))
        maxdd = (
            pd.DataFrame(parts, columns=["_dow", "_hour", "maxdd_pct"])
            .pivot(index="_dow", columns="_hour", values="maxdd_pct")
        )

        trade_count.to_csv(dirs["tables"] / "heatmap_tradeCount.csv")
        win_rate_mat.to_csv(dirs["tables"] / "heatmap_winRate_pct.csv")
        pnl_sum.to_csv(dirs["tables"] / "heatmap_pnl_sum.csv")
        netpct_mat.to_csv(dirs["tables"] / "heatmap_netpct.csv")
        maxdd.to_csv(dirs["tables"] / "heatmap_maxdd_pct.csv")

        plot_heatmap(trade_count, dirs["heatmaps"] / "heatmap_tradeCount.png",
                     "Trade Count — Day vs Hour (Filtered)", fmt=".0f", higher_is_better=True)
        plot_heatmap(win_rate_mat, dirs["heatmaps"] / "heatmap_winRate_pct.png",
                     "Win Rate (%) — Day vs Hour (Filtered)", fmt=".1f", higher_is_better=True)
        plot_heatmap(pnl_sum, dirs["heatmaps"] / "heatmap_pnl_sum.png",
                     "PnL Sum (cash) — Day vs Hour (Filtered)", fmt=".0f", higher_is_better=True)
        plot_heatmap(netpct_mat, dirs["heatmaps"] / "heatmap_netpct.png",
                     "Net PnL (%) — Day vs Hour (Filtered)", fmt=".2f", higher_is_better=True)
        plot_heatmap(maxdd, dirs["heatmaps"] / "heatmap_maxdd_pct.png",
                     "Max Drawdown (%) — Day vs Hour (Filtered)", fmt=".2f", higher_is_better=False)

    print("[OK] News-filtered report generated.")
    print(f"Run folder : {run_dir}")
    print(f"Summary    : {dirs['meta'] / 'summary.json'}")
    print(f"XLSX       : {xlsx_path}")
    print(f"Figures    : {dirs['figures']}")
    print(f"Figure (by strat): {dirs['figures'] / 'equity_drawdown_pct_by_strategy.png'}")
    print(f"Heatmaps   : {dirs['heatmaps']}")
    print(f"Redacted   : {dirs['tables'] / 'redacted_trades.csv'}")


if __name__ == "__main__":
    main()
