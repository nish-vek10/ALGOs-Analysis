# src/scripts/01_quant_report_processed.py
"""
Quant-style report from a PROCESSED closed-trades CSV (CLOSED POSITIONS ONLY).

Assumption (CONFIRMED BY YOU):
- Column 'Profit' is PER-TRADE realized profit (NOT cumulative).
- Ignore commission (do not subtract it).

Reads:
  data/processed/<...>__offset-2h.csv

Writes (no overlaps):
  output/runs/<RUN_ID>/
    meta/summary.json
    tables/trades_with_equity.csv
    tables/heatmap_*.csv
    figures/equity_drawdown_pct.png
    figures/heatmaps/heatmap_*.png
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# =========================
# CONFIG
# =========================
PROCESSED_CSV = Path("data/processed/ReportHistory-52651106__acct-52651106__offset-2h.csv")

START_EQUITY = 10000.0

MT_TIME_FMT = "%Y.%m.%d %H:%M:%S"  # processed timestamps are kept in this MT style

PNL_COL = "Profit"  # per-trade PnL column (ignore commission)

# Breakeven handling:
DROP_ZERO_PNL_ROWS = False
ZERO_EPS = 1e-9

# Strategy name normalization (from Comment column)
STRATEGY_COL = "Comment"
STRATEGY_MAP = {
    "Inv_ATRtrail_pen": "ATR-Trailing STRAT",
    "Inv_Chandelier_p": "Chandelier STRAT",
}
STRATEGY_ORDER = ["ATR-Trailing STRAT", "Chandelier STRAT"]


# =========================
# PROJECT PATHS
# =========================
def project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def extract_account_id_from_filename(p: Path) -> str:
    m = re.search(r"(\d{5,})", p.stem)
    return m.group(1) if m else "unknown"


def safe_slug(s: str, max_len: int = 90) -> str:
    s = re.sub(r"[^A-Za-z0-9_\-]+", "_", s).strip("_")
    return s[:max_len] if len(s) > max_len else s


def utc_run_id(acct: str, stem: str) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") + "Z"
    return f"{ts}__acct-{safe_slug(acct, 40)}__{safe_slug(stem, 80)}"


def ensure_run_dirs(run_dir: Path) -> Dict[str, Path]:
    dirs = {
        "run": run_dir,
        "meta": run_dir / "meta",
        "tables": run_dir / "tables",
        "figures": run_dir / "figures",
        "heatmaps": run_dir / "figures" / "heatmaps",
        "report": run_dir / "report",
    }
    for p in dirs.values():
        p.mkdir(parents=True, exist_ok=True)
    return dirs


# =========================
# COLUMN DETECTION
# =========================
def find_time_column(df: pd.DataFrame, kind: str) -> Optional[str]:
    if kind not in {"open", "close"}:
        raise ValueError("kind must be 'open' or 'close'")

    candidates = (
        ["OpenTime", "Open Time", "open_time", "opentime"] if kind == "open"
        else ["CloseTime", "Close Time", "close_time", "closetime"]
    )

    cols_lower = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in cols_lower:
            return cols_lower[c.lower()]

    # soft fallback
    for c in df.columns:
        cl = c.lower()
        if kind == "open" and ("open" in cl and "time" in cl):
            return c
        if kind == "close" and ("close" in cl and "time" in cl):
            return c

    return None


def normalize_strategy(raw: object) -> str:
    s = "" if raw is None else str(raw).strip()
    return STRATEGY_MAP.get(s, s if s else "Unknown STRAT")


def coerce_numeric(series: pd.Series) -> pd.Series:
    """
    Robust number parsing for MT-style exports where:
      - thousands separator can be a SPACE: '2 467.84'
      - negatives can be written with a space after minus: '- 999.02'
      - may include non-breaking spaces
    """
    s = series.astype(str)

    # normalize unicode minus + remove NBSP
    s = s.str.replace("\u2212", "-", regex=False).str.replace("\xa0", "", regex=False)

    # remove spaces (both thousands separators and '- 999.02' style)
    s = s.str.replace(" ", "", regex=False)

    # remove commas if any appear
    s = s.str.replace(",", "", regex=False)

    # keep only digits, dot, minus
    s = s.str.replace(r"[^0-9\.\-]", "", regex=True)

    return pd.to_numeric(s, errors="coerce")


# =========================
# TIME PARSING
# =========================
def parse_mt_time(series: pd.Series) -> pd.Series:
    dt = pd.to_datetime(series, format=MT_TIME_FMT, errors="coerce")
    if dt.notna().sum() == 0:
        dt = pd.to_datetime(series, errors="coerce")
    return dt


# =========================
# METRICS
# =========================
def max_drawdown_pct(equity: np.ndarray, start_equity: float) -> Tuple[float, int, int]:
    """
    Max drawdown as % of start_equity, based on equity curve.
    Returns (max_dd_pct, peak_idx, trough_idx)
    """
    if len(equity) == 0 or start_equity == 0:
        return 0.0, -1, -1
    peaks = np.maximum.accumulate(equity)
    dd_cash = peaks - equity
    trough_idx = int(np.argmax(dd_cash))
    peak_idx = int(np.argmax(peaks[:trough_idx + 1])) if trough_idx >= 0 else 0
    max_dd_cash = float(dd_cash[trough_idx]) if len(dd_cash) else 0.0
    return float((max_dd_cash / start_equity) * 100.0), peak_idx, trough_idx


def streaks(per_trade_pnl: np.ndarray) -> Dict[str, float]:
    if len(per_trade_pnl) == 0:
        return {
            "max_win_streak_n": 0,
            "max_loss_streak_n": 0,
            "max_win_streak_cash": 0.0,
            "max_loss_streak_cash": 0.0,
        }

    signs = np.where(per_trade_pnl > 0, 1, np.where(per_trade_pnl < 0, -1, 0))

    max_win_n = 0
    max_loss_n = 0
    max_win_cash = 0.0
    max_loss_cash = 0.0

    cur_sign = 0
    cur_n = 0
    cur_cash = 0.0

    for p, s in zip(per_trade_pnl, signs):
        if s == 0:
            cur_sign = 0
            cur_n = 0
            cur_cash = 0.0
            continue

        if s == cur_sign:
            cur_n += 1
            cur_cash += float(p)
        else:
            cur_sign = int(s)
            cur_n = 1
            cur_cash = float(p)

        if cur_sign == 1:
            max_win_n = max(max_win_n, cur_n)
            max_win_cash = max(max_win_cash, cur_cash)
        else:
            max_loss_n = max(max_loss_n, cur_n)
            max_loss_cash = min(max_loss_cash, cur_cash)

    return {
        "max_win_streak_n": int(max_win_n),
        "max_loss_streak_n": int(max_loss_n),
        "max_win_streak_cash": float(max_win_cash),
        "max_loss_streak_cash": float(max_loss_cash),
    }


def streaks_metrics(df_sub: pd.DataFrame) -> dict:
    pnl = df_sub["_pnl"].to_numpy(dtype=float) if len(df_sub) else np.array([], dtype=float)
    s = streaks(pnl)
    return {
        "max_win_streak_n": int(s["max_win_streak_n"]),
        "max_loss_streak_n": int(s["max_loss_streak_n"]),
        "max_win_streak_cash": float(s["max_win_streak_cash"]),
        "max_loss_streak_cash": float(s["max_loss_streak_cash"]),
    }


def sharpe_ratio_per_trade(pnl: np.ndarray, start_equity: float) -> float:
    """
    Simple per-trade Sharpe using returns = pnl / start_equity.
    This is not annualized (no time scaling), it's a clean trade-to-trade quality metric.
    """
    if start_equity <= 0 or len(pnl) < 2:
        return 0.0
    rets = pnl / float(start_equity)
    mu = float(np.mean(rets))
    sd = float(np.std(rets, ddof=1))
    if sd == 0.0:
        return 0.0
    return mu / sd


def compute_metrics(df_sub: pd.DataFrame, start_equity: float) -> dict:
    """
    Returns a totals dict with the SAME fields as your current overall totals,
    plus sharpe_ratio and recovery_factor.
    """
    total_trades = int(len(df_sub))
    pnl = df_sub["_pnl"].to_numpy(dtype=float) if total_trades else np.array([], dtype=float)

    winners = int((df_sub["_pnl"] > 0).sum())
    losers = int((df_sub["_pnl"] < 0).sum())
    breakeven = int((df_sub["_pnl"] == 0.0).sum())

    win_rate = (winners / total_trades) * 100.0 if total_trades else 0.0

    cum_pnl = np.cumsum(pnl) if total_trades else np.array([], dtype=float)
    equity = start_equity + cum_pnl if total_trades else np.array([start_equity], dtype=float)

    total_pnl = float(cum_pnl[-1]) if total_trades else 0.0
    end_equity = float(equity[-1]) if total_trades else float(start_equity)
    net_pct = ((end_equity / float(start_equity)) - 1.0) * 100.0 if start_equity else 0.0

    max_dd_pct_pos, _, _ = max_drawdown_pct(equity, start_equity)

    gross_win = float(df_sub.loc[df_sub["_pnl"] > 0, "_pnl"].sum()) if winners else 0.0
    gross_loss = float(df_sub.loc[df_sub["_pnl"] < 0, "_pnl"].sum()) if losers else 0.0  # negative
    profit_factor = float(gross_win / abs(gross_loss)) if losers else float("inf")

    expectancy = float(total_pnl / total_trades) if total_trades else 0.0

    best_trade = float(df_sub["_pnl"].max()) if total_trades else 0.0
    worst_trade = float(df_sub["_pnl"].min()) if total_trades else 0.0

    avg_win = float(df_sub.loc[df_sub["_pnl"] > 0, "_pnl"].mean()) if winners else 0.0
    avg_loss = float(df_sub.loc[df_sub["_pnl"] < 0, "_pnl"].mean()) if losers else 0.0

    # Hold mins
    hold_valid = df_sub["_hold_minutes"].notna().any()
    avg_hold = float(df_sub["_hold_minutes"].mean()) if hold_valid else None
    med_hold = float(df_sub["_hold_minutes"].median()) if hold_valid else None
    p95_hold = float(df_sub["_hold_minutes"].quantile(0.95)) if hold_valid else None

    # Sharpe + Recovery
    sharpe = float(sharpe_ratio_per_trade(pnl, start_equity)) if total_trades else 0.0
    recovery = float(net_pct / max_dd_pct_pos) if max_dd_pct_pos not in (0.0, -0.0) else 0.0

    return {
        "totalTrades": total_trades,
        "winners": winners,
        "losers": losers,
        "breakeven": breakeven,
        "winRate_pct": win_rate,

        "startEquity": float(start_equity),
        "endEquity": end_equity,
        "pnl_cash": total_pnl,
        "netPct": net_pct,

        "maxDD_pct": max_dd_pct_pos,
        "grossWin": gross_win,
        "grossLoss": gross_loss,
        "profitFactor": profit_factor,

        "expectancy_per_trade": expectancy,
        "bestTrade": best_trade,
        "worstTrade": worst_trade,
        "avgWin": avg_win,
        "avgLoss": avg_loss,

        "avgHoldMin": avg_hold,
        "medianHoldMin": med_hold,
        "p95HoldMin": p95_hold,

        "sharpe_ratio": sharpe,
        "recovery_factor": recovery,
    }


# =========================
# PLOTS
# =========================
def plot_equity_dd(
    times: pd.Series,
    equity: np.ndarray,
    start_equity: float,
    out_path: Path,
    title: str,
) -> None:
    """
    Soft theme + presentation-friendly styling (matplotlib only):
    - Light canvas background
    - White "card" axes
    - Equity % line + filled from 0% baseline (green above, red below)
    - Drawdown % line in red + filled
    """
    import matplotlib.dates as mdates

    times = pd.to_datetime(times)
    equity = np.asarray(equity, dtype=float)

    # Equity %
    eq_pct = ((equity / float(start_equity)) - 1.0) * 100.0 if start_equity else np.zeros_like(equity)

    # Drawdown %
    peaks = np.maximum.accumulate(equity)
    dd_cash = peaks - equity
    dd_pct = (dd_cash / float(start_equity)) * 100.0 if start_equity else np.zeros_like(dd_cash)
    dd_pct = -dd_pct  # negative values

    # ---- Figure canvas (soft background) ----
    fig = plt.figure(figsize=(16, 7.5), facecolor="#F5F7FB", constrained_layout=True)
    gs = fig.add_gridspec(2, 1, height_ratios=[2.5, 1.15])

    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1], sharex=ax1)

    # ---- Axis "card" style ----
    def style_card(ax):
        ax.set_facecolor("white")
        ax.grid(True, which="major", alpha=0.18, linewidth=1)
        ax.grid(True, which="minor", alpha=0.08, linewidth=0.8)
        for side in ["top", "right"]:
            ax.spines[side].set_visible(False)
        for side in ["left", "bottom"]:
            ax.spines[side].set_alpha(0.25)
        ax.tick_params(axis="both", labelsize=10)

    style_card(ax1)
    style_card(ax2)

    # ---- Title ----
    ax1.set_title(title, fontsize=16, fontweight="bold", pad=12)

    # ---- Equity line + fill from 0 baseline ----
    ax1.plot(times, eq_pct, linewidth=2.6, alpha=0.95)
    ax1.axhline(0, linewidth=1.0, alpha=0.35)

    # Fill green above 0, red below 0 (same baseline)
    ax1.fill_between(times, eq_pct, 0, where=(eq_pct >= 0), alpha=0.18, interpolate=True)
    ax1.fill_between(times, eq_pct, 0, where=(eq_pct < 0), color="red", alpha=0.10, interpolate=True)

    ax1.set_ylabel("Equity (%)", fontsize=12)
    ax1.tick_params(axis="x", labelbottom=False)

    # ---- Drawdown line in red + fill ----
    ax2.plot(times, dd_pct, linewidth=2.2, color="red", alpha=0.95)
    ax2.fill_between(times, dd_pct, 0, color="red", alpha=0.12)
    ax2.axhline(0, linewidth=1.0, alpha=0.35)

    ax2.set_ylabel("Drawdown (%)", fontsize=12)
    ax2.set_xlabel("Time", fontsize=12)

    # Drawdown limits (padding)
    dd_min = float(np.nanmin(dd_pct)) if len(dd_pct) else 0.0
    ax2.set_ylim(min(dd_min * 1.15, -0.5), 0.0)

    # ---- Date formatting ----
    ax2.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=10))
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    ax2.xaxis.set_minor_locator(mdates.AutoDateLocator(minticks=12, maxticks=20))

    # Make date labels clean
    fig.autofmt_xdate(rotation=0)

    # ---- Optional end annotation ----
    if len(eq_pct):
        ax1.annotate(
            f"End: {eq_pct[-1]:.2f}%",
            xy=(times.iloc[-1], eq_pct[-1]),
            xytext=(-12, 10),
            textcoords="offset points",
            ha="right",
            fontsize=10,
            alpha=0.85,
        )

    fig.savefig(out_path, dpi=240, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_heatmap(mat: pd.DataFrame, out_path: Path, title: str, fmt: str, higher_is_better: bool) -> None:
    """
    Day (Mon..Sun) x Hour (0..23) annotated heatmap.
    RdYlGn: red worse, green better.
    If lower is better, use reversed cmap so green = low.
    """
    day_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    mat = mat.reindex(day_order)
    mat = mat.reindex(columns=list(range(24)))

    values = mat.to_numpy(dtype=float)
    cmap = plt.cm.RdYlGn if higher_is_better else plt.cm.RdYlGn_r

    fig, ax = plt.subplots(figsize=(16, 5.5), constrained_layout=True)
    im = ax.imshow(values, aspect="auto", cmap=cmap)

    ax.set_title(title)
    ax.set_xlabel("Hour of day (chart time)")
    ax.set_ylabel("Day of week")

    ax.set_xticks(np.arange(24))
    ax.set_xticklabels([str(h) for h in range(24)])
    ax.set_yticks(np.arange(len(day_order)))
    ax.set_yticklabels(day_order)

    cbar = fig.colorbar(im, ax=ax, fraction=0.02, pad=0.02)
    cbar.ax.tick_params(labelsize=9)

    for i in range(values.shape[0]):
        for j in range(values.shape[1]):
            v = values[i, j]
            if np.isfinite(v):
                ax.text(j, i, format(v, fmt), ha="center", va="center", fontsize=7)

    fig.savefig(out_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def bin_max_dd_pct(sub: pd.DataFrame, start_equity: float) -> float:
    """
    Max drawdown within a bin, computed on the bin's own cumulative PnL path,
    scaled by start_equity for comparability.
    """
    if len(sub) == 0 or start_equity == 0:
        return np.nan
    sub = sub.sort_values("_open_dt")
    cum = np.cumsum(sub["_pnl"].to_numpy(dtype=float))
    peaks = np.maximum.accumulate(cum)
    dd = peaks - cum
    return float(np.max(dd) / float(start_equity) * 100.0) if len(dd) else 0.0


# =========================
# EXCEL SUMMARY
# =========================
def write_summary_xlsx(path: Path, rows: list[dict]) -> None:
    """
    rows: list of dicts with keys used below.
    Writes a simple CEO-friendly table with formatting.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Strategy",
        "totalTrades", "winners", "losers", "breakeven",
        "startEquity", "endEquity",
        "winRate_pct",
        "pnl_cash", "netPct",
        "maxDD_pct",
        "profitFactor",
        "expectancy_per_trade",
        "bestTrade", "worstTrade",
        "avgWin", "avgLoss",
        "grossWin", "grossLoss",
        "sharpe_ratio", "recovery_factor",
        "avgHoldMin", "medianHoldMin", "p95HoldMin",
    ]

    ws.append(headers)

    # header style
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align

    # data rows
    for r in rows:
        ws.append([r.get(h) for h in headers])

    # column widths (reasonable)
    widths = {
        "A": 22, "B": 12, "C": 10, "D": 10, "E": 12,
        "F": 14, "G": 14, "H": 12, "I": 12,
        "J": 14, "K": 18, "L": 12, "M": 14, "N": 12,
        "O": 14, "P": 16, "Q": 14, "R": 14
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # number formats:
    # - 4dp for these fields in xlsx (as requested)
    four_dp_cols = {
        "winRate_pct", "netPct", "maxDD_pct", "profitFactor",
        "expectancy_per_trade", "avgHoldMin", "medianHoldMin", "p95HoldMin",
        "sharpe_ratio", "recovery_factor"
    }

    # - money fields 2dp
    two_dp_cols = {
        "pnl_cash", "grossWin", "grossLoss",
        "bestTrade", "worstTrade",
        "avgWin", "avgLoss",
        "startEquity", "endEquity",
        "max_win_streak_cash", "max_loss_streak_cash",
    }

    # map header name to column index
    header_to_idx = {h: i + 1 for i, h in enumerate(headers)}

    for row_idx in range(2, 2 + len(rows)):
        for h, col_idx in header_to_idx.items():
            cell = ws.cell(row=row_idx, column=col_idx)

            # Align numbers nicely, keep Strategy left
            if h == "Strategy":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                continue
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if h in {"totalTrades", "winners", "losers", "breakeven", "max_win_streak_n", "max_loss_streak_n"}:
                cell.number_format = "0"
            elif h in four_dp_cols:
                cell.number_format = "0.0000"
            elif h in two_dp_cols:
                cell.number_format = "0.00"

    wb.save(path)


# =========================
# MAIN
# =========================
def main() -> None:
    root = project_root()
    processed_path = (root / PROCESSED_CSV).resolve()
    if not processed_path.exists():
        raise FileNotFoundError(f"Processed CSV not found: {processed_path}")

    df = pd.read_csv(processed_path)

    open_col = find_time_column(df, "open")
    close_col = find_time_column(df, "close")
    if open_col is None:
        raise ValueError(f"Could not find an OpenTime column. Columns={list(df.columns)}")

    if PNL_COL not in df.columns:
        raise ValueError(f"Expected per-trade PnL column '{PNL_COL}' not found. Columns={list(df.columns)}")

    df["_open_dt"] = parse_mt_time(df[open_col])
    df["_close_dt"] = parse_mt_time(df[close_col]) if close_col is not None else pd.NaT

    # Closed-only enforcement (protective)
    df = df[df["_open_dt"].notna()].copy()
    if close_col is not None:
        df = df[df["_close_dt"].notna()].copy()

    # Per-trade PnL from Profit column (ignore commission)
    df["_pnl"] = coerce_numeric(df[PNL_COL]).fillna(0.0)

    # Normalize strategy name from Comment column
    if STRATEGY_COL in df.columns:
        df["_strategy"] = df[STRATEGY_COL].apply(normalize_strategy)
    else:
        df["_strategy"] = "Unknown STRAT"

    # Remove near-zero rows if requested (to enforce breakeven=0)
    if DROP_ZERO_PNL_ROWS:
        df = df[df["_pnl"].abs() > ZERO_EPS].copy()

    df.sort_values("_open_dt", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Hold time metrics (minutes)
    if close_col is not None:
        df["_hold_minutes"] = (df["_close_dt"] - df["_open_dt"]).dt.total_seconds() / 60.0
    else:
        df["_hold_minutes"] = np.nan

    acct = extract_account_id_from_filename(processed_path)
    run_id = utc_run_id(acct=acct, stem=processed_path.stem)
    run_dir = (root / "output" / "runs" / run_id).resolve()
    dirs = ensure_run_dirs(run_dir)

    # Equity curve (account equity from start balance + cumulative PnL)
    pnl = df["_pnl"].to_numpy(dtype=float)
    cum_pnl = np.cumsum(pnl)
    equity = START_EQUITY + cum_pnl

    # Drawdown %
    peaks = np.maximum.accumulate(equity)
    dd_cash = peaks - equity
    dd_pct = (dd_cash / float(START_EQUITY)) * 100.0 if START_EQUITY else np.zeros_like(dd_cash)
    dd_pct = -dd_pct

    # =========================
    # SUMMARY METRICS (overall)
    # =========================
    overall_metrics = compute_metrics(df, START_EQUITY)
    total_trades = overall_metrics["totalTrades"]

    # Keep streaks (your existing streak function)
    st = streaks(pnl)

    # Keep max dd window indices for timestamps
    max_dd_pct_pos, peak_i, trough_i = max_drawdown_pct(equity, START_EQUITY)

    summary = {
        "input": {
            "processed_csv": str(processed_path),
            "open_time_col": open_col,
            "close_time_col": close_col,
            "pnl_col": PNL_COL,
            "start_equity": START_EQUITY,
            "time_format": MT_TIME_FMT,
            "drop_zero_pnl_rows": DROP_ZERO_PNL_ROWS,
            "zero_eps": ZERO_EPS,
            "commission_ignored": True,
        },
        "totals": overall_metrics,
        "streaks": st,
        "max_dd_window": {
            "peak_index": int(peak_i),
            "trough_index": int(trough_i),
            "peak_time_open": str(df["_open_dt"].iloc[peak_i]) if peak_i >= 0 and total_trades else None,
            "trough_time_open": str(df["_open_dt"].iloc[trough_i]) if trough_i >= 0 and total_trades else None,
        },
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "run_dir": str(run_dir),
    }

    # =========================
    # PER-STRATEGY METRICS
    # =========================
    by_strategy = {}
    for strat_name, df_s in df.groupby("_strategy", dropna=False):
        by_strategy[strat_name] = compute_metrics(df_s, START_EQUITY)

    summary["by_strategy"] = by_strategy

    (dirs["meta"] / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    # =========================
    # CEO XLSX SUMMARY (2 strats + overall)
    # =========================
    rows_for_xlsx = []

    for strat in STRATEGY_ORDER:
        df_s = df[df["_strategy"] == strat]
        m = compute_metrics(df_s, START_EQUITY)
        m.update(streaks_metrics(df_s))
        m["Strategy"] = strat
        rows_for_xlsx.append(m)

    overall_row = overall_metrics.copy()
    overall_row.update(streaks_metrics(df))
    overall_row["Strategy"] = "OVERALL (PORTFOLIO)"
    rows_for_xlsx.append(overall_row)

    xlsx_path = dirs["report"] / "summary.xlsx"
    write_summary_xlsx(xlsx_path, rows_for_xlsx)

    # Save trades with equity & dd
    df_out = df.copy()
    df_out["CumPnL"] = cum_pnl
    df_out["Equity"] = equity
    df_out["Drawdown_pct"] = dd_pct
    df_out.to_csv(dirs["tables"] / "trades_with_equity.csv", index=False)

    # Plot equity + drawdown
    plot_equity_dd(
        times=df["_open_dt"],
        equity=equity,
        start_equity=START_EQUITY,
        out_path=dirs["figures"] / "equity_drawdown_pct.png",
        title=f"Equity Curve & Drawdown (pct) — acct {acct}",
    )

    # Heatmaps by OPEN time
    df["_dow"] = df["_open_dt"].dt.day_name().str[:3]
    df["_hour"] = df["_open_dt"].dt.hour
    df["_is_win"] = (df["_pnl"] > 0).astype(int)

    trade_count = df.pivot_table(index="_dow", columns="_hour", values="_pnl", aggfunc="count")
    win_rate_mat = df.pivot_table(index="_dow", columns="_hour", values="_is_win", aggfunc="mean") * 100.0
    pnl_sum = df.pivot_table(index="_dow", columns="_hour", values="_pnl", aggfunc="sum")
    netpct_mat = (pnl_sum / float(START_EQUITY) * 100.0) if START_EQUITY else pnl_sum * 0.0

    parts = []
    for (dow, hour), sub in df[["_dow", "_hour", "_open_dt", "_pnl"]].groupby(["_dow", "_hour"], sort=False):
        parts.append((dow, hour, bin_max_dd_pct(sub, START_EQUITY)))
    maxdd = (
        pd.DataFrame(parts, columns=["_dow", "_hour", "maxdd_pct"])
        .pivot(index="_dow", columns="_hour", values="maxdd_pct")
    )

    # Save matrices
    trade_count.to_csv(dirs["tables"] / "heatmap_tradeCount.csv")
    win_rate_mat.to_csv(dirs["tables"] / "heatmap_winRate_pct.csv")
    pnl_sum.to_csv(dirs["tables"] / "heatmap_pnl_sum.csv")
    netpct_mat.to_csv(dirs["tables"] / "heatmap_netpct.csv")
    maxdd.to_csv(dirs["tables"] / "heatmap_maxdd_pct.csv")

    # Plot heatmaps (RdYlGn: red worse -> green better)
    plot_heatmap(
        trade_count,
        dirs["heatmaps"] / "heatmap_tradeCount.png",
        "Trade Count — Day vs Hour (Open Time, chart time)",
        fmt=".0f",
        higher_is_better=True,
    )
    plot_heatmap(
        win_rate_mat,
        dirs["heatmaps"] / "heatmap_winRate_pct.png",
        "Win Rate (%) — Day vs Hour (Open Time, chart time)",
        fmt=".1f",
        higher_is_better=True,
    )
    plot_heatmap(
        pnl_sum,
        dirs["heatmaps"] / "heatmap_pnl_sum.png",
        "PnL Sum (cash) — Day vs Hour (Open Time, chart time)",
        fmt=".0f",
        higher_is_better=True,
    )
    plot_heatmap(
        netpct_mat,
        dirs["heatmaps"] / "heatmap_netpct.png",
        "Net PnL (%) — Day vs Hour (Open Time, chart time)",
        fmt=".2f",
        higher_is_better=True,
    )
    plot_heatmap(
        maxdd,
        dirs["heatmaps"] / "heatmap_maxdd_pct.png",
        "Max Drawdown (%) — Day vs Hour (within bin)",
        fmt=".2f",
        higher_is_better=False,  # lower DD is better => green
    )

    print("[OK] Quant report generated.")
    print(f"Run folder : {run_dir}")
    print(f"Summary    : {dirs['meta'] / 'summary.json'}")
    print(f"Figures    : {dirs['figures']}")
    print(f"Heatmaps   : {dirs['heatmaps']}")


if __name__ == "__main__":
    main()
